"""Admin routes: dashboard, matching, overrides, analytics, and export."""

import io
from datetime import datetime, timezone

from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, session, send_file
from flask_login import login_required, current_user
from sqlalchemy.orm import joinedload

from models import db, Student, Guide, Preference, Allocation, Notification, AuditLog
from matching import run_matching
from utils import log_action, role_required

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')


@admin_bp.route('/dashboard')
@login_required
@role_required('admin')
def dashboard():
    students = Student.query.options(joinedload(Student.user)).all()
    guides = Guide.query.options(joinedload(Guide.user)).all()
    allocations = Allocation.query.all()
    unmatched = [s for s in students if not Allocation.query.filter_by(student_id=s.id).first()]
    notifications = (Notification.query
                     .filter_by(user_id=current_user.id)
                     .order_by(Notification.created_at.desc())
                     .limit(10).all())
    audit_logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).limit(20).all()

    total_capacity = sum(g.capacity for g in guides)
    total_allocated = sum(1 for a in allocations if a.status in ('allocated', 'confirmed'))

    stats = {
        'total_students': len(students),
        'total_guides': len(guides),
        'total_allocated': total_allocated,
        'total_unmatched': len(unmatched),
        'total_capacity': total_capacity,
        'completion_pct': round((total_allocated / len(students) * 100) if students else 0, 1),
        'pref_submitted': Preference.query.count(),
    }

    # Guide load data for charts
    guide_load_data = []
    for g in guides:
        guide_load_data.append({
            'name': g.user.name,
            'capacity': g.capacity,
            'load': g.current_load,
            'areas': ', '.join(g.research_areas[:3])
        })

    # Preference distribution
    pref_dist = {'choice_1': 0, 'choice_2': 0, 'choice_3': 0, 'fallback': 0}
    for a in allocations:
        if a.preference_rank == 1:
            pref_dist['choice_1'] += 1
        elif a.preference_rank == 2:
            pref_dist['choice_2'] += 1
        elif a.preference_rank == 3:
            pref_dist['choice_3'] += 1
        else:
            pref_dist['fallback'] += 1

    return render_template('admin/dashboard.html',
                           students=students, guides=guides,
                           allocations=allocations, unmatched=unmatched,
                           stats=stats, guide_load_data=guide_load_data,
                           pref_dist=pref_dist, notifications=notifications,
                           audit_logs=audit_logs)


@admin_bp.route('/run-matching', methods=['POST'])
@login_required
@role_required('admin')
def run_matching_route():
    result = run_matching()
    flash(f'Matching complete! Phase 1: {result["phase1_matched"]}, Phase 2: {result["phase2_matched"]}, '
          f'Unmatched: {result["unmatched"]}', 'success')
    session['matching_result'] = result
    return redirect(url_for('admin.dashboard'))


@admin_bp.route('/override', methods=['POST'])
@login_required
@role_required('admin')
def override():
    try:
        student_id = int(request.form.get('student_id'))
        guide_id = int(request.form.get('guide_id'))
    except (TypeError, ValueError):
        flash('Invalid student or guide ID.', 'error')
        return redirect(url_for('admin.dashboard'))

    # Remove existing allocation if any
    existing = Allocation.query.filter_by(student_id=student_id).first()
    if existing:
        old_guide = db.session.get(Guide, existing.guide_id)
        if old_guide:
            old_guide.current_load = max(0, old_guide.current_load - 1)
        db.session.delete(existing)

    # Create new allocation
    guide = db.session.get(Guide, guide_id)
    student = db.session.get(Student, student_id)
    alloc = Allocation(
        student_id=student_id,
        guide_id=guide_id,
        status='allocated',
        method='manual',
        preference_rank=0,
        allocated_at=datetime.now(timezone.utc)
    )
    db.session.add(alloc)
    guide.current_load += 1

    # Notify
    notif = Notification(
        user_id=student.user_id, type='info', title='Guide Reassigned',
        message=f'You have been manually assigned to {guide.user.name} by admin.'
    )
    db.session.add(notif)
    db.session.commit()
    log_action('admin_override', target=f'student={student_id}', details=f'assigned to guide={guide_id}')
    flash(f'Manually assigned {student.user.name} → {guide.user.name}', 'success')
    return redirect(url_for('admin.dashboard'))


@admin_bp.route('/analytics')
@login_required
@role_required('admin')
def analytics():
    guides = Guide.query.options(joinedload(Guide.user)).all()
    allocations = Allocation.query.all()
    students = Student.query.all()

    # Most preferred guides
    guide_demand = {}
    for g in guides:
        count = Preference.query.filter(
            (Preference.choice_1_id == g.id) |
            (Preference.choice_2_id == g.id) |
            (Preference.choice_3_id == g.id)
        ).count()
        guide_demand[g.user.name] = count

    # Method distribution
    method_dist = {}
    for a in allocations:
        method_dist[a.method] = method_dist.get(a.method, 0) + 1

    # CGPA distribution of allocated students
    cgpa_ranges = {'Below 7': 0, '7-8': 0, '8-9': 0, '9+': 0}
    for s in students:
        if s.cgpa < 7:
            cgpa_ranges['Below 7'] += 1
        elif s.cgpa < 8:
            cgpa_ranges['7-8'] += 1
        elif s.cgpa < 9:
            cgpa_ranges['8-9'] += 1
        else:
            cgpa_ranges['9+'] += 1

    return render_template('admin/analytics.html',
                           guides=guides, guide_demand=guide_demand,
                           method_dist=method_dist, cgpa_ranges=cgpa_ranges,
                           allocations=allocations, students=students)


@admin_bp.route('/export')
@login_required
@role_required('admin')
def export():
    """Export allocations to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Allocations"

        # Header styling
        header_fill = PatternFill(start_color="6C63FF", end_color="6C63FF", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = ['S.No', 'Student Name', 'Email', 'CGPA', 'Department',
                   'Guide Name', 'Status', 'Method', 'Preference Rank', 'Allocated At']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        allocations = Allocation.query.all()

        # Pre-fetch students and guides for faster lookups
        students_map = {s.id: s for s in Student.query.options(joinedload(Student.user)).all()}
        guides_map = {g.id: g for g in Guide.query.options(joinedload(Guide.user)).all()}

        for row_idx, alloc in enumerate(allocations, 2):
            student = students_map.get(alloc.student_id)
            guide = guides_map.get(alloc.guide_id)
            data = [
                row_idx - 1,
                student.user.name if student else 'N/A',
                student.user.email if student else 'N/A',
                student.cgpa if student else 0,
                student.user.department if student else 'N/A',
                guide.user.name if guide else 'N/A',
                alloc.status,
                alloc.method,
                f'Choice #{alloc.preference_rank}' if alloc.preference_rank > 0 else 'Fallback',
                alloc.allocated_at.strftime('%Y-%m-%d %H:%M') if alloc.allocated_at else 'N/A'
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

        # Auto-width
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col) + 2
            ws.column_dimensions[col[0].column_letter].width = max_length

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name='guide_allocation_report.xlsx'
        )
    except ImportError:
        flash('openpyxl is required for Excel export. Install it with: pip install openpyxl', 'error')
        return redirect(url_for('admin.dashboard'))
